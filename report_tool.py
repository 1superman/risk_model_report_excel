from openpyxl.styles import Font, colors, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from sklearn.metrics import roc_auc_score,roc_curve
from matplotlib.font_manager import FontProperties
from scipy.stats import ks_2samp
from woe_iv import cut_bin, woe_picture, get_auto_bin
import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
from matplotlib import pylab
from sklearn import metrics
import math
import joblib
import re
import os

_TONT = FontProperties(fname=r"c:\windows\fonts\simsun.ttc", size=14)
__FONT1 = Font(name='Times New Roman', size=11, color=colors.BLACK, bold=True)
__FONT2 = Font(name='Times New Roman', size=11, color=colors.BLACK, bold=False)
_ALIGNMENT1 = Alignment(horizontal='left', vertical='center')
_ALIGNMENT2 = Alignment(horizontal='center', vertical='center')

def get_ks(y_true, y_pred): 
    if y_pred[y_true==1].shape[0]==0 or y_pred[y_true!=1].shape[0]==0:
        return np.nan
    else:    
        return round(ks_2samp(y_pred[y_true==1], y_pred[y_true!=1]).statistic,3)

def get_roc_auc_score(y_true,y_pred):
    if y_true.nunique()!=2:
        return np.nan
    else :
        return round(roc_auc_score(y_true,y_pred),3)

def get_num(string):
    return float(re.findall("[0-9.]+", str(string))[1])

def to_score(x):
    if x <=0.001:
        x =0.001
    elif x >=0.999:
        x =0.999
    A = 404.65547022
    B = 72.1347520444
    result = round(int(A-B*math.log(x/(1-x))))
    
    if result < 300:
        result=300
    if result>900:
        result=900
    return result

def plot_roc(df_src_with_all):
    plt.figure(figsize=(4,4))
    lw = 1
    for splitted_type, df_splitted_type in df_src_with_all.groupby('type'):
        fpr, tpr, thresholds = roc_curve(df_splitted_type['y_label'], df_splitted_type['p'])
        plt.plot(fpr, tpr,lw=lw, label=splitted_type+'_curve (area = %0.4f)' % metrics.auc(fpr,tpr))

    plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel('False Positive Rate')
    plt.ylabel('True Positive Rate')
    plt.title('ALL_TR_TE_OOT ROC')
    plt.legend(loc="lower right")
    plt.savefig('ALL_TR_TE_OOT ROC', bbox_inches='tight')

def plot_ks_(df):
    def plot_ks(y_true: pd.Series, y_pred: pd.Series, output_path, title=''):
    	fpr, tpr, thresholds = metrics.roc_curve(y_true, y_pred)
    	plt.figure(figsize=(4, 4))
    	x = np.arange(len(thresholds))
    	x_show = np.arange(0, len(thresholds), int(len(thresholds) / 10))
    	pylab.plot(x, tpr, lw=1)
    	pylab.plot(x, fpr, lw=1)
    	pylab.plot(x, tpr - fpr, lw=1, linestyle='--')
    	pylab.xticks(x_show, [thresholds[i] for i in x_show], rotation=90)
    	pylab.xlabel('score')
    	pylab.title(title)
    	plt.savefig(output_path, bbox_inches='tight')
    	return
    for type_, df_ in df.groupby('type'):
        title = type_+' KS'
        plot_ks(df_['y_label'],df_['p'], output_path=title,title=title)             
    return

def plot_tot_describe(df_src_with_all):
    font = _TONT
    for splitted_type, df_splitted_type in df_src_with_all.groupby('type'):
        plt.figure(figsize=(6,4))
        sns.distplot( df_splitted_type[df_splitted_type['type']==splitted_type]["score"] , color='#ff8080',bins=20,kde_kws={ "lw": 2.5, 'linestyle':'--'})
        tem = df_splitted_type[df_splitted_type['type']==splitted_type]["score"].describe().reset_index()
        table_ = plt.table(cellText = [[round(x,4)] for x in tem['score'].tolist()],
                  colWidths=[0.1]*1,rowLabels = tem['index'].tolist(),loc='right')
        table_.set_fontsize(15)
        table_.scale(1.9,2.265)
        plt.title('{}评分分布'.format(splitted_type), fontproperties=font)
        plt.savefig('{}评分分布'.format(splitted_type), bbox_inches='tight')
    color_list = ['#d65a31','#40bfc1','#b7e778','#5edfff']
    plt.figure(figsize=(6,4))
    a = -1
    for splitted_type, df_splitted_type in df_src_with_all.groupby('type'):
        a = a+1
        sns.kdeplot(df_splitted_type[df_splitted_type['type']==splitted_type]["score"] ,shade_lowest=False,color=color_list[a],
                    label = splitted_type)
    plt.title('各样本的评分分布对比', fontproperties=font)
    plt.savefig('各样本的评分分布对比', bbox_inches='tight')
    return

def single_type_distribution(df, asc=1):
    df_ = df.groupby('cut').apply(lambda item: pd.Series({
                                                        'total_cnt': item.shape[0],
                                                        'bad_cnt': item[item['y_label'] == 1].shape[0],
                                                        'good_cnt': item[item['y_label'] == 0].shape[0]}))\
                                                        [['total_cnt', 'bad_cnt', 'good_cnt']]
    df_ = df_.sort_index(ascending=asc)
    df_['bad_prob'] = df_['bad_cnt'] / df_['bad_cnt'].sum()
    df_['good_prob'] = df_['good_cnt'] / df_['good_cnt'].sum()
    df_['cum_bad_prob'] = df_['bad_prob'].cumsum()
    df_['cum_good_prob'] = df_['good_prob'].cumsum()
    df_['bad_rate'] = df_['bad_cnt'] / df_['total_cnt']
    df_['cum_total_cnt'] = df_['total_cnt'].cumsum()
    df_['cum_bad_cnt'] = df_['bad_cnt'].cumsum()
    df_['cum_pass'] = df_['cum_total_cnt'] / df_['total_cnt'].sum()
    df_['cum_pass_bad_rate'] = df_['cum_bad_cnt'] / df_['cum_total_cnt']
    df_.index.name = 'score_bin'
    df_ = df_.reset_index()

    df_str_distruibution = df_.copy()
    int_cols = ['total_cnt', 'bad_cnt', 'good_cnt', 'cum_total_cnt', 'cum_bad_cnt']
    df_str_distruibution[int_cols] = df_str_distruibution[int_cols].applymap(
        lambda v: str(int(v)) if pd.notnull(v) else '')
    rate_cols = ['bad_prob', 'good_prob', 'cum_bad_prob', 'cum_good_prob', 'bad_rate', 'cum_pass',
                 'cum_pass_bad_rate']
    df_str_distruibution[rate_cols] = df_str_distruibution[rate_cols]\
                                        .applymap(lambda v: '{:.1%}'.format(round(v, 3)) \
                                                  if pd.notnull(v) else '')
    df_str_distruibution = df_str_distruibution[
                          ['score_bin', 'total_cnt', 'bad_cnt', 'good_cnt', 'bad_prob', 'good_prob', 'cum_bad_prob', 
                           'cum_good_prob', 'bad_rate', 'cum_total_cnt', 'cum_bad_cnt', 'cum_pass', 'cum_pass_bad_rate']]
    df_str_distruibution.columns = ['评分', '组内总人数', '组内坏客户数', '组内好客户数', '坏客户占比', '好客户占比', 
                                    '累计坏客户占比', '累计好客户占比', '区间违约率', '累计拒绝人数', '累计拒绝坏人数', '累计拒绝率', '累计拒绝坏人占比']
    return df_str_distruibution

def write_xlsx(df, sheet, index=3, pace=7):
    for i in range(0, df.shape[1]):
        col = get_column_letter(index+i)
        values = df.columns[i]
        sheet[col+str(pace)].value = values
        sheet[col+str(pace)].font = __FONT1
        sheet[col+str(pace)].alignment = _ALIGNMENT1
        len_base = len(str(values))
        base_values = values
        for j in range(0, df.shape[0]):
            if df.iloc[j, i] in ['type', 'index']:
                continue
            values = df.iloc[j, i]
            if df.columns[i] in ['feature', 'psi', 'iv_sum']:
                if (values == base_values):
                    sheet[col+str(pace+j+1)].value = ''
                else:
                    sheet[col+str(pace+j+1)].value = values
                    base_values = values
            else:
                sheet[col+str(pace+j+1)].value = values
            sheet[col+str(pace+j+1)].font = __FONT2
            sheet[col+str(pace+j+1)].alignment = _ALIGNMENT1
            len_base = max(len(str(values)), len_base)
        if len_base >= 50:
            sheet.column_dimensions[col].width = 50
        elif len_base >= 30:
            sheet.column_dimensions[col].width = 30
        elif len_base >= 10:
            sheet.column_dimensions[col].width = 10
        else:
            continue
    return sheet

#一. Overall
def Overall(wb, df):
    sheet = wb.create_sheet('Overall',0)
    df1 = df.groupby(['type', df['triger_date'].str[:7]])['y_label'].agg(["count","mean"])\
                .reset_index().rename(columns={"triger_date": u"月份", "sum":u"人数", "mean":"逾期率"})

    df2 = pd.concat([df.groupby(['type', df['triger_date'].str[:7]]).apply(lambda item: get_ks(item['y_label'], item['p'])), \
                     df.groupby(['type', df['triger_date'].str[:7]]).apply(lambda item: get_roc_auc_score(item['y_label'], item['p']))], axis=1) \
            .reset_index().rename(columns={"triger_date": u"月份", 0: "KS", 1: "AUC"})

    sheet = write_xlsx(df1, sheet, index=3, pace=7)
    sheet = write_xlsx(df2, sheet, index=3, pace=16)
    return wb

#二. feature_importance
def Feature_importance(wb, df, online_feature):
    sheet = wb.create_sheet('mode_report',1)

    online_feature = online_feature[['feature_name', 'descreption']]
    feature_importance = pd.DataFrame(data=ml._Booster.get_fscore().items(), columns=['feature_name', 'feature_importance'])\
                            .merge(online_feature, on='feature_name', how='left')
                            
    feature_importance = feature_importance.sort_values('feature_importance', ascending=False, ignore_index=True)
    sheet = write_xlsx(feature_importance, sheet, index=1, pace=2)

    df['score'] = df['p'].map(to_score)
    df['cut'] = pd.qcut(df['score'], q=10, duplicates='drop')
    df['cut'] = df['cut'].map(get_num)

    plot_roc(df)
    plot_ks_(df)
    plot_tot_describe(df)

    df_distribution_test = single_type_distribution(df.loc[df['type'] == 'test'])
    df_distribution_oot = single_type_distribution(df.loc[df['type'] == 'oot'])

    sheet.merge_cells("E3:Q3")
    sheet['E3'].value = u"时间内测试集"
    sheet['E3'].font = __FONT1
    sheet['E3'].alignment = _ALIGNMENT2
    sheet = write_xlsx(df_distribution_test, sheet, index=5, pace=4)

    sheet.merge_cells("E16:Q16")
    sheet['E16'].value = u"时间外测试集"
    sheet['E16'].font = __FONT1
    sheet['E16'].alignment = _ALIGNMENT2

    sheet = write_xlsx(df_distribution_oot, sheet, index=5, pace=17)
    sheet.column_dimensions['D'].width = 50
    img=Image('ALL_TR_TE_OOT ROC.png')
    newSize=(300,300)
    img.width,img.height=newSize
    sheet.add_image(img,"D30")

    img=Image('train KS.png')
    newSize=(300,300)
    img.width,img.height=newSize
    sheet.add_image(img,"D48")

    img=Image('test KS.png')
    newSize=(300,300)
    img.width,img.height=newSize
    sheet.add_image(img,"E48")

    img=Image('oot KS.png')
    newSize=(300,300)
    img.width,img.height=newSize
    sheet.add_image(img,"J48")

    img=Image('oot KS.png')
    newSize=(300,300)
    img.width,img.height=newSize
    sheet.add_image(img,"J48")

    img=Image('各样本的评分分布对比.png')
    newSize=(300,300)
    img.width,img.height=newSize
    sheet.add_image(img,"D66")
    return wb

# 三. feature_describe
def Feature_describe(wb, df, ml):
    sheet = wb.create_sheet('feature_describe',2)

    sheet.merge_cells("A1:Q1")
    sheet['A1'].value = u"模型入模特征统计描述"
    sheet['A1'].font = __FONT1
    sheet['A1'].alignment = _ALIGNMENT2
    df_ = df[ml._Booster.feature_names].describe([i/10 for i in range(10)]).T.reset_index()
    df_['missing_rate'] = 1-df_['count']/df.shape[0]
    df_[df_.columns[1:]] = df_[df_.columns[1:]].applymap(lambda x:round(x,4))
    sheet = write_xlsx(df_, sheet, index=1, pace=2)
    return wb

#四. WOE_IV
def Woe_IV(wb, df, ml, df_bins):
    sheet = wb.create_sheet('WOE_IV',3)
    df_bins = df_bins.loc[df_bins['字段英文名'].isin(ml._Booster.feature_names)]
    df_woe1, df_woe2 = cut_bin(df, df_bins)
    ivandwoelist1, ivandwoelist2 = woe_picture(df_woe2, df_bins, excel_name='woe_plot', y='y_label', flag=1, \
                                               feature='字段英文名', binx='binx', woex='woe', delimiter=',')
        
    sheet.merge_cells("A1:N1")
    sheet['A1'].value = u"时间内样本训练集WOE分布"
    sheet['A1'].font = __FONT1
    sheet['A1'].alignment = _ALIGNMENT2
    sheet = write_xlsx(ivandwoelist1, sheet, index=1, pace=2)

    sheet.merge_cells("Q1:AC1")
    sheet['Q1'].value = u"时间内样本测试集WOE分布"
    sheet['Q1'].font = __FONT1
    sheet['Q1'].alignment = _ALIGNMENT2
    sheet = write_xlsx(ivandwoelist2, sheet, index=17, pace=2)
    return wb

if __name__ == "__main__":
    ml = joblib.load("xgbmodel.ml")
    df = pd.read_pickle("result.pkl")
    try:
        df_bins = pd.read_csv('woe_bin.csv')
    except:   
        df_bins = get_auto_bin(df.loc[df['type'] == 'train'], ml._Booster.feature_names)
    online_feature = pd.read_excel("线上特征.xlsx")
    online_feature = online_feature.drop_duplicates('feature_name', ignore_index=True)
    try:
        os.mkdir('2020Q4现金贷(有宏伟特征)')
    except:
        pass
    os.chdir('2020Q4现金贷(有宏伟特征)')
    wb = Workbook()
    wb = Overall(wb, df)
    wb = Feature_importance(wb, df, online_feature)
    wb = Feature_describe(wb, df, ml)
    wb = Woe_IV(wb, df, ml, df_bins)
    wb.save(u'id_pdlold_117报告.xlsx')
    

        